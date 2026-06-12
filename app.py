"""
The Circle Visitor Reconciliation Tool
=======================================
Accepts one Excel workbook covering an entire month:
  Sheet 1 – Visitor Report    (one row per visitor per day)
  Sheet 2 – Card Register     (normalised: Date | Card Number | Name | Company | Returned)
  Sheet 3 – Access Log        (Name="N VISITOR" | Date | First Swipe | Last Swipe)

Matching keys use (Date + Name) so the same card number used on different days
is matched correctly to each day's visitor.

Output workbook:
  - Final Report  (fully populated Card No., Check-In, Check-Out, Duration)
  - Discrepancies (unmatched visitors, missing check-outs, duplicate entries)
  - Dashboard     (KPI summary, Top Hosts, Top Companies, unreturned cards)

Branding  →  Yellow #FFCD33 | Black #000000 | White #FFFFFF
"""

import difflib
import io
import re
from datetime import datetime, timedelta

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTS — exact sheet & column names as found in the workbook
# ──────────────────────────────────────────────────────────────────────────────

SHEET_VISITOR = "Visitor Report"
SHEET_CARD    = "Card Register"     # renamed from "Card Allocation"; now normalised with Date column
SHEET_ACCESS  = "Access Log"
REQUIRED_SHEETS = [SHEET_VISITOR, SHEET_CARD, SHEET_ACCESS]

# Visitor Report columns
VR_DATE     = "Date"
VR_CARD     = "Card No."
VR_CHECKIN  = "Check-In Time"
VR_CHECKOUT = "Check Out Time"
VR_DURATION = "Duration"
VR_VISITOR  = "Visitor Name"
VR_EMAIL    = "Email ID"
VR_HOST     = "Host Name"
VR_COMPANY  = "Company Name"
VR_PURPOSE  = "Purpose of Visit"
REQUIRED_VR_COLS = [VR_DATE, VR_CARD, VR_CHECKIN, VR_CHECKOUT, VR_DURATION,
                    VR_VISITOR, VR_EMAIL, VR_HOST, VR_COMPANY, VR_PURPOSE]

# Card Register columns — normalised format now includes Date
# Matching key: (Date + Name) → card number
CA_DATE     = "Date"          # NEW: date the card was issued
CA_CARDNUM  = "Card Number"   # was "Card number"
CA_NAME     = "Name"
CA_COMPANY  = "Company"
CA_RETURNED = "Returned"
REQUIRED_CA_COLS = [CA_DATE, CA_CARDNUM, CA_NAME, CA_COMPANY, CA_RETURNED]

# Access Log columns
# Matching key: (Date + Card Number) — card number extracted from "N VISITOR"
AL_NAME  = "Name"
AL_DATE  = "Date"
AL_FIRST = "First Swipe"
AL_LAST  = "Last Swipe"       # capitalised to match new format
REQUIRED_AL_COLS = [AL_NAME, AL_DATE, AL_FIRST, AL_LAST]

# Brand colours (hex without #)
YELLOW  = "FFCD33"
BLACK   = "000000"
WHITE   = "FFFFFF"
LIGHT_Y = "FFF8DC"   # cornsilk — alternating row shading
RED_LT  = "FFE0E0"   # light red — discrepancy category highlight


# ──────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG & CSS
# ──────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="The Circle – Visitor Reconciliation",
    page_icon="🟡",
    layout="wide",
)

st.markdown(
    """
    <style>
      [data-testid="stAppViewContainer"] { background-color: #FFFFFF; }

      /* ── branded header ── */
      .tc-header {
          background-color: #FFCD33;
          padding: 22px 32px 16px 32px;
          border-radius: 10px;
          margin-bottom: 24px;
      }
      .tc-header h1 { color:#000; font-size:2rem; font-weight:800; margin:0; }
      .tc-header p  { color:#333; font-size:0.95rem; margin:6px 0 0 0; }

      /* ── KPI metric cards ── */
      .metric-card {
          background:#FFCD33; border-radius:8px; padding:16px 22px;
          text-align:center; margin-bottom:8px;
      }
      .metric-card .val { font-size:2.4rem; font-weight:900; color:#000; line-height:1; }
      .metric-card .lbl {
          font-size:0.75rem; font-weight:700; color:#333;
          margin-top:4px; text-transform:uppercase; letter-spacing:0.5px;
      }

      /* ── section sub-headers ── */
      .sec-head {
          font-size:1.05rem; font-weight:700; color:#000;
          border-left:5px solid #FFCD33; padding-left:10px;
          margin:20px 0 10px 0;
      }

      /* ── Generate Report button ── */
      [data-testid="stButton"] > button {
          background-color:#000 !important; color:#FFCD33 !important;
          font-weight:700 !important; border-radius:6px !important;
          border:none !important; padding:10px 28px !important;
      }
      [data-testid="stButton"] > button:hover { background-color:#222 !important; }

      /* ── Download button ── */
      [data-testid="stDownloadButton"] button {
          background-color:#FFCD33 !important; color:#000 !important;
          font-weight:700 !important; border:2px solid #000 !important;
          border-radius:6px !important; padding:10px 28px !important;
      }
      [data-testid="stDownloadButton"] button:hover { background-color:#e6b800 !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

# ──────────────────────────────────────────────────────────────────────────────
# HEADER
# ──────────────────────────────────────────────────────────────────────────────

st.markdown(
    """
    <div class="tc-header">
        <h1>🟡 The Circle Visitor Reconciliation Tool</h1>
        <p>Upload the monthly visitor workbook → reconcile cards &amp; access logs by date → download the completed report.</p>
    </div>
    """,
    unsafe_allow_html=True,
)


# ──────────────────────────────────────────────────────────────────────────────
# UTILITY FUNCTIONS
# ──────────────────────────────────────────────────────────────────────────────

def fmt_time(val) -> str:
    """
    Normalise any time-like value to "HH:MM". Returns "" for NaN / None /
    unparseable input.

    Handles:
      - timedelta           e.g. Timedelta('0 days 05:49:00')
      - datetime / Timestamp e.g. Timestamp('2026-06-10 05:49:00')
      - "HH:MM" / "HH:MM:SS" strings
      - Excel serial time as float (fraction of a day, e.g. 0.243055..)
    """
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass  # pd.isna can choke on some exotic types — fall through

    if isinstance(val, timedelta):
        total_mins = int(val.total_seconds() // 60)
        return f"{total_mins // 60:02d}:{total_mins % 60:02d}"

    if isinstance(val, datetime):
        return val.strftime("%H:%M")

    # Excel sometimes stores times as a bare float = fraction of a 24h day
    if isinstance(val, (int, float)):
        try:
            total_mins = int(round(float(val) * 24 * 60))
            total_mins %= (24 * 60)
            return f"{total_mins // 60:02d}:{total_mins % 60:02d}"
        except (TypeError, ValueError):
            return ""

    # String forms: "05:49:00", "05:49", "5:49 AM", etc.
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return ""

    # Try strict HH:MM[:SS] first
    parts = s.split(":")
    if len(parts) >= 2:
        try:
            return f"{int(parts[0]) % 24:02d}:{int(parts[1]):02d}"
        except ValueError:
            pass

    # Fall back to pandas' generic time parser (handles "5:49 AM" etc.)
    try:
        t = pd.to_datetime(s)
        return t.strftime("%H:%M")
    except Exception:
        return ""


def calc_duration(checkin: str, checkout: str) -> str:
    """
    Compute elapsed time as "HH:MM".
    Returns "" if either argument is missing, empty, or unparseable.
    """
    if not checkin or not checkout:
        return ""
    if not isinstance(checkin, str) or not isinstance(checkout, str):
        return ""
    try:
        t_in  = datetime.strptime(checkin,  "%H:%M")
        t_out = datetime.strptime(checkout, "%H:%M")
        delta = t_out - t_in
        if delta.total_seconds() < 0:       # handles rare past-midnight case
            delta += timedelta(hours=24)
        total_mins = int(delta.total_seconds() // 60)
        return f"{total_mins // 60:02d}:{total_mins % 60:02d}"
    except (ValueError, TypeError):
        return ""


def _normalise(text: str) -> str:
    """Lowercase + collapse whitespace."""
    return re.sub(r"\s+", " ", str(text).strip().lower())


def _split_tokens(text: str) -> list[str]:
    """
    Split card-name / visitor-name on spaces AND on delimiter chars (+ & / -).
    Filters out tokens shorter than 3 characters to avoid noise.
    Returns [] for blank/NaN-like input ("", "nan", "none", etc.).
    """
    norm = _normalise(text)
    if norm in ("", "nan", "none", "nat"):
        return []
    cleaned = re.sub(r"[+&/\-]", " ", norm)
    return [t for t in cleaned.split() if len(t) >= 3]


def smart_match(card_name: str, visitor_name: str) -> bool:
    """
    Multi-strategy matching:

    For every token in card_name, attempt (in order):
      1. Exact token match against any visitor_name token.
      2. Fuzzy similarity ≥ 0.80 (handles typos like purvi/purvy, sweta/shweta).
         A length-ratio guard (visitor token ≥ 75 % of card token length) prevents
         short visitor tokens (e.g. "Anush") from false-matching longer card names
         (e.g. "anushka").
      3. Substring containment — card token appears literally inside visitor_name.

    All card_name tokens must match; any failing token means no overall match.
    """
    cn_tokens = _split_tokens(card_name)
    vn_tokens = _split_tokens(visitor_name)
    vn_full   = _normalise(visitor_name)

    if not cn_tokens:
        return False

    for ct in cn_tokens:
        matched = False

        # Strategy 1 & 2: compare against every visitor token
        for vt in vn_tokens:
            # length-ratio guard: skip if visitor token is too short vs card token
            if len(ct) > 0 and (len(vt) / len(ct)) < 0.75:
                continue
            similarity = difflib.SequenceMatcher(None, ct, vt).ratio()
            if similarity >= 0.80:
                matched = True
                break

        # Strategy 3: substring fallback (e.g. "joseph" inside "Joseph + Scott…")
        if not matched and ct in vn_full:
            matched = True

        if not matched:
            return False  # this card token couldn't be paired → no match

    return True


def find_card_for_visitor(
    visit_date: str,
    visitor_name: str,
    card_df: pd.DataFrame,
) -> tuple[int | None, str | None]:
    """
    Search Card Register for a row whose (Date + Name) matches
    (visit_date + visitor_name).

    Matching order:
      1. Date must match exactly (normalised string comparison).
      2. Name: exact match first, then smart_match (fuzzy + substring).

    Returns (card_number, matched_card_name) or (None, None).
    """
    if card_df.empty or "_date_key" not in card_df.columns:
        return None, None

    date_key = _normalise(str(visit_date))

    # Pre-filter to the same date for speed — avoids comparing names cross-day
    day_cards = card_df[card_df["_date_key"] == date_key]

    for _, row in day_cards.iterrows():
        card_name = row[CA_NAME]
        if pd.isna(card_name) or str(card_name).strip() == "":
            continue
        if smart_match(str(card_name), visitor_name):
            return int(row[CA_CARDNUM]), str(card_name).strip()

    return None, None


def find_access_record(
    visit_date: str,
    card_no: int,
    access_df: pd.DataFrame,
) -> "pd.Series | None":
    """
    Locate the Access Log row matching (Date + Card Number).

    The Name column contains values like "15 VISITOR"; the card number is
    extracted from the prefix.  Date is matched as a normalised string so that
    different Excel date representations (Timestamp, string) all compare equal.

    Returns the first matching row Series, or None.
    """
    if access_df.empty or "_date_key" not in access_df.columns:
        return None

    date_key = _normalise(str(visit_date))
    target   = f"{card_no} VISITOR"

    matches = access_df[
        (access_df["_date_key"] == date_key) &
        (access_df[AL_NAME].astype(str).str.strip() == target)
    ]
    return None if matches.empty else matches.iloc[0]


# ──────────────────────────────────────────────────────────────────────────────
# WORKBOOK VALIDATION
# ──────────────────────────────────────────────────────────────────────────────

def _norm_header(name: str) -> str:
    """Normalise a column header for comparison: lowercase, collapse whitespace."""
    return re.sub(r"\s+", " ", str(name).strip().lower())


def _norm_sheet_name(name: str) -> str:
    """Normalise a sheet name for comparison: lowercase, collapse whitespace."""
    return re.sub(r"\s+", " ", str(name).strip().lower())


def validate_workbook(xl: pd.ExcelFile) -> list[str]:
    """
    Check that all required sheets and columns exist.

    Sheet and column matching is case-insensitive and whitespace-tolerant
    (e.g. "card register", "Card  Register", " Card Register " all match
    "Card Register"). Returns a list of human-readable error strings;
    empty list = valid.
    """
    errors: list[str] = []

    # 1. Verify sheet presence (case/whitespace-insensitive)
    actual_sheets_norm = {_norm_sheet_name(s): s for s in xl.sheet_names}
    for sheet in REQUIRED_SHEETS:
        if _norm_sheet_name(sheet) not in actual_sheets_norm:
            errors.append(
                f"Missing sheet: '{sheet}' "
                f"(found sheets: {', '.join(xl.sheet_names)})"
            )

    if errors:          # can't check columns if sheets are absent
        return errors

    # 2. Verify column presence (case/whitespace-insensitive)
    col_checks = [
        (SHEET_VISITOR, REQUIRED_VR_COLS),
        (SHEET_CARD,    REQUIRED_CA_COLS),
        (SHEET_ACCESS,  REQUIRED_AL_COLS),
    ]
    for sheet_name, required_cols in col_checks:
        actual_sheet_name = actual_sheets_norm[_norm_sheet_name(sheet_name)]
        df     = pd.read_excel(xl, sheet_name=actual_sheet_name, nrows=0)
        actual_norm = {_norm_header(c): str(c).strip() for c in df.columns}
        for col in required_cols:
            if _norm_header(col) not in actual_norm:
                errors.append(
                    f"Sheet '{sheet_name}' — missing column: '{col}' "
                    f"(found columns: {', '.join(str(c).strip() for c in df.columns)})"
                )

    return errors


# ──────────────────────────────────────────────────────────────────────────────
# CORE RECONCILIATION ENGINE
# ──────────────────────────────────────────────────────────────────────────────

def reconcile(xl: pd.ExcelFile) -> dict:
    """
    Full multi-date reconciliation workflow.

    Matching keys:
      Step 1-2 : (Date + Visitor Name)  →  Card Register  →  Card Number
      Step 3-4 : (Date + Card Number)   →  Access Log     →  First/Last Swipe
      Step 5   : Duration = Last Swipe − First Swipe (HH:MM)

    Returns a dict with keys:
        report_df       – completed Visitor Report DataFrame
        discrepancies   – DataFrame of all flagged issues
        dashboard       – dict of KPI name → value (includes Top Hosts / Companies)
        unreturned_df   – Card Register rows where card was not returned
    """

    # ── 1. Load sheets (case/whitespace-insensitive sheet name lookup) ───────
    actual_sheets_norm = {_norm_sheet_name(s): s for s in xl.sheet_names}
    visitor_sheet = actual_sheets_norm[_norm_sheet_name(SHEET_VISITOR)]
    card_sheet    = actual_sheets_norm[_norm_sheet_name(SHEET_CARD)]
    access_sheet  = actual_sheets_norm[_norm_sheet_name(SHEET_ACCESS)]

    visitor_df = pd.read_excel(xl, sheet_name=visitor_sheet)
    card_df    = pd.read_excel(xl, sheet_name=card_sheet)
    access_df  = pd.read_excel(xl, sheet_name=access_sheet)

    # Normalise column headers to the canonical names this app expects,
    # tolerating case/whitespace differences (e.g. "card  number" -> "Card Number")
    def _rename_to_canonical(df: pd.DataFrame, canonical_cols: list[str]) -> pd.DataFrame:
        rename_map = {}
        canon_by_norm = {_norm_header(c): c for c in canonical_cols}
        for col in df.columns:
            norm = _norm_header(col)
            if norm in canon_by_norm:
                rename_map[col] = canon_by_norm[norm]
            else:
                rename_map[col] = str(col).strip()
        return df.rename(columns=rename_map)

    visitor_df = _rename_to_canonical(visitor_df, REQUIRED_VR_COLS)
    card_df    = _rename_to_canonical(card_df,    REQUIRED_CA_COLS)
    access_df  = _rename_to_canonical(access_df,  REQUIRED_AL_COLS)

    # ── Clean Card Register ───────────────────────────────────────────────────
    # Drop rows with no card number or no name
    # Card Number may arrive as int, float (1.0), or numeric string ("1");
    # use pd.to_numeric with errors="coerce" so all of these parse correctly
    # while non-numeric junk (NaN, text) becomes NaN and gets filtered out.
    _card_num_numeric = pd.to_numeric(card_df[CA_CARDNUM], errors="coerce")
    card_df = card_df[
        card_df[CA_NAME].notna() &
        (card_df[CA_NAME].astype(str).str.strip() != "") &
        _card_num_numeric.notna()
    ].copy()
    card_df[CA_CARDNUM] = _card_num_numeric[card_df.index].astype(int)

    # Pre-compute a normalised date key on each sheet for fast lookup.
    #
    # The three sheets use three different date representations:
    #   Visitor Report : Timestamp with time component  e.g. 2026-05-01 16:00:00
    #   Card Register  : string with no year            e.g. "22-May" or "5-May"
    #   Access Log     : Timestamp date-only            e.g. 2026-05-01 00:00:00
    #
    # Normalise everything to zero-padded "MM-DD" so all three line up.

    def _to_mmdd(v) -> str:
        """
        Convert any date representation to a zero-padded "MM-DD" string.

        Handles, in order:
          1. NaN / None                        -> ""
          2. pandas Timestamp / datetime        -> direct month/day
          3. Excel serial date as a bare number -> convert via pandas origin
          4. "22-May" / "5-May" (no year)        -> assume current/parsed year
          5. Any other string pandas can parse  -> generic to_datetime
          6. Total failure                      -> lowercase string fallback
             (keeps the pipeline running rather than crashing; will simply
             fail to match, which surfaces as an "Unmatched Visitor" row
             that's easy to spot in Discrepancies)
        """
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except (TypeError, ValueError):
            pass

        # Timestamp / datetime -> extract month and day directly
        if isinstance(v, (pd.Timestamp, datetime)):
            return f"{v.month:02d}-{v.day:02d}"

        # Excel serial date number (e.g. 46000.0) — pandas epoch is 1899-12-30
        if isinstance(v, (int, float)):
            try:
                t = pd.Timestamp("1899-12-30") + pd.Timedelta(days=float(v))
                return f"{t.month:02d}-{t.day:02d}"
            except Exception:
                return ""

        s = str(v).strip()
        if not s or s.lower() == "nan":
            return ""

        # "22-May" or "5-May" format (no year) — attach a placeholder year
        # so pandas can parse month/day; the year itself is irrelevant since
        # we only ever compare MM-DD across sheets within the same workbook.
        try:
            t = pd.Timestamp(f"{s}-2026")
            return f"{t.month:02d}-{t.day:02d}"
        except Exception:
            pass

        # Generic fallback — handles "10/06/2026", "2026-06-10", "10 Jun 2026", etc.
        try:
            t = pd.to_datetime(s, dayfirst=True, errors="raise")
            return f"{t.month:02d}-{t.day:02d}"
        except Exception:
            pass

        # Last resort: return a lowercase string so two identical "weird"
        # values still match each other even if we can't parse a real date
        return s.lower()

    card_df["_date_key"]   = card_df[CA_DATE].apply(_to_mmdd)
    access_df["_date_key"] = access_df[AL_DATE].apply(_to_mmdd)

    # Normalise Access Log Name column
    access_df[AL_NAME] = access_df[AL_NAME].astype(str).str.strip()

    # ── 2-5. Row-by-row reconciliation ───────────────────────────────────────
    report_rows:     list[dict] = []
    unmatched:       list[dict] = []   # no card found in Card Register
    card_not_in_log: list[dict] = []   # card found but absent from Access Log
    missing_co:      list[dict] = []   # check-in present, check-out blank
    # Duplicate key: (date, normalised visitor name) so same person on different
    # days is NOT counted as a duplicate
    duplicate_check: dict = {}

    for idx, vrow in visitor_df.iterrows():
        raw_name = vrow[VR_VISITOR]

        # Skip completely blank rows (no visitor name) — these are usually
        # stray empty rows at the end of a sheet and shouldn't be reported
        # as "unmatched visitors"
        if pd.isna(raw_name) or str(raw_name).strip() == "":
            continue

        visitor_name = str(raw_name).strip()
        visit_date   = vrow[VR_DATE]

        # Normalise visit_date using the same MM-DD key as Card Register / Access Log
        date_key = _to_mmdd(visit_date) if not isinstance(visit_date, str) else _to_mmdd(visit_date)

        # Track duplicates per (date, visitor) — same person on same day
        dup_key = (date_key, _normalise(visitor_name))
        duplicate_check.setdefault(dup_key, []).append(idx)

        # STEP 1 & 2 — match (Date + Visitor Name) → Card Register → Card Number
        card_no, matched_card_name = find_card_for_visitor(
            date_key, visitor_name, card_df
        )

        check_in = check_out = duration = ""

        if card_no is not None:
            # STEP 3 & 4 — match (Date + Card Number) → Access Log → swipes
            acc_row = find_access_record(date_key, card_no, access_df)

            if acc_row is None:
                card_not_in_log.append({
                    "Date"        : visit_date,
                    "Visitor Name": visitor_name,
                    "Card No."    : card_no,
                    "Issue"       : f"Card {card_no} not found in Access Log for {date_key}",
                })
            else:
                check_in  = fmt_time(acc_row[AL_FIRST])
                check_out = fmt_time(acc_row[AL_LAST])
                # STEP 5 — compute duration
                duration  = calc_duration(check_in, check_out)

                if not check_out:
                    missing_co.append({
                        "Date"         : visit_date,
                        "Visitor Name" : visitor_name,
                        "Card No."     : card_no,
                        "Check-In Time": check_in,
                        "Issue"        : "Missing check-out in Access Log",
                    })
        else:
            unmatched.append({
                "Date"        : visit_date,
                "Visitor Name": visitor_name,
                "Email ID"    : vrow.get(VR_EMAIL, ""),
                "Host Name"   : vrow.get(VR_HOST,  ""),
                "Issue"       : "No matching card in Card Register for this date",
            })

        # Build final report row (preserving original column order)
        report_rows.append({
            VR_DATE    : visit_date,
            VR_CARD    : card_no if card_no is not None else "",
            VR_CHECKIN : check_in,
            VR_CHECKOUT: check_out,
            VR_DURATION: duration,
            VR_VISITOR : visitor_name,
            VR_EMAIL   : vrow.get(VR_EMAIL,   ""),
            VR_HOST    : vrow.get(VR_HOST,    ""),
            VR_COMPANY : vrow.get(VR_COMPANY, ""),
            VR_PURPOSE : vrow.get(VR_PURPOSE, ""),
        })

    # ── Detect same-day duplicate visitors ───────────────────────────────────
    duplicate_visitors: list[dict] = []
    for (d_key, vn_key), indices in duplicate_check.items():
        if len(indices) > 1:
            display_name = visitor_df.iloc[indices[0]][VR_VISITOR]
            display_date = visitor_df.iloc[indices[0]][VR_DATE]
            excel_rows   = ", ".join(str(i + 2) for i in indices)
            duplicate_visitors.append({
                "Date"        : display_date,
                "Visitor Name": display_name,
                "Issue"       : (
                    f"Visitor appears {len(indices)} times on this date "
                    f"(rows {excel_rows})"
                ),
            })

    # ── Build Discrepancies DataFrame ─────────────────────────────────────────
    disc_rows: list[dict] = []

    for r in unmatched:
        disc_rows.append({
            "Category"    : "Unmatched Visitor",
            "Date"        : r["Date"],
            "Visitor Name": r["Visitor Name"],
            "Card No."    : "",
            "Detail"      : r["Issue"],
        })
    for r in card_not_in_log:
        disc_rows.append({
            "Category"    : "Card Not in Access Log",
            "Date"        : r["Date"],
            "Visitor Name": r["Visitor Name"],
            "Card No."    : r["Card No."],
            "Detail"      : r["Issue"],
        })
    for r in missing_co:
        disc_rows.append({
            "Category"    : "Missing Check-Out",
            "Date"        : r["Date"],
            "Visitor Name": r["Visitor Name"],
            "Card No."    : r["Card No."],
            "Detail"      : r["Issue"],
        })
    for r in duplicate_visitors:
        disc_rows.append({
            "Category"    : "Duplicate Visitor",
            "Date"        : r["Date"],
            "Visitor Name": r["Visitor Name"],
            "Card No."    : "",
            "Detail"      : r["Issue"],
        })

    disc_df = pd.DataFrame(disc_rows) if disc_rows else pd.DataFrame(
        columns=["Category", "Date", "Visitor Name", "Card No.", "Detail"]
    )

    # ── Unreturned cards ──────────────────────────────────────────────────────
    # "Returned"-type columns can come in many shapes across files:
    #   "Returned" (text)    -> returned
    #   "B"                  -> borrowed / not returned
    #   "Yes" / "Y" / True   -> returned
    #   "No" / "N" / False   -> not returned
    #   blank / NaN          -> not returned (assume still out)
    #
    # We treat anything that looks like an affirmative "returned" signal as
    # returned, and everything else (including unrecognised text) as
    # unreturned — erring on the side of flagging more cards for review
    # rather than silently hiding them.
    RETURNED_POSITIVE = {"RETURNED", "YES", "Y", "TRUE", "DONE", "OK", "RETD", "RET"}

    def _is_returned(v) -> bool:
        if v is None:
            return False
        try:
            if pd.isna(v):
                return False
        except (TypeError, ValueError):
            pass
        if isinstance(v, bool):
            return v is True
        s = str(v).strip().upper()
        return s in RETURNED_POSITIVE

    unreturned_mask = ~card_df[CA_RETURNED].apply(_is_returned)
    unreturned_df = card_df[
        unreturned_mask & card_df[CA_NAME].notna()
    ].copy()

    # ── Dashboard KPIs ────────────────────────────────────────────────────────
    # Build with explicit columns so an empty result still has the right
    # shape (avoids KeyError on report_df[VR_CARD] etc. when the Visitor
    # Report sheet had zero usable rows)
    report_columns = [VR_DATE, VR_CARD, VR_CHECKIN, VR_CHECKOUT, VR_DURATION,
                      VR_VISITOR, VR_EMAIL, VR_HOST, VR_COMPANY, VR_PURPOSE]
    report_df = pd.DataFrame(report_rows, columns=report_columns)

    total_visitors  = len(report_df)
    matched_count   = int((report_df[VR_CARD] != "").sum()) if total_visitors else 0

    # Unique visitors: distinct (date, normalised name) pairs
    if total_visitors:
        unique_visitors = report_df.apply(
            lambda r: (
                _normalise(str(r[VR_DATE])),
                _normalise(str(r[VR_VISITOR]))
            ),
            axis=1,
        ).nunique()
    else:
        unique_visitors = 0

    # Top 5 Hosts (by visit count)
    top_hosts = (
        report_df[VR_HOST]
        .dropna()
        .pipe(lambda s: s[s.astype(str).str.strip() != ""])
        .value_counts()
        .head(5)
        .to_dict()
    ) if total_visitors else {}

    # Top 5 Companies (by visit count)
    top_companies = (
        report_df[VR_COMPANY]
        .dropna()
        .pipe(lambda s: s[s.astype(str).str.strip() != ""])
        .value_counts()
        .head(5)
        .to_dict()
    ) if total_visitors else {}

    dashboard = {
        "Total Visitors"    : total_visitors,
        "Unique Visitors"   : unique_visitors,
        "Matched Visitors"  : matched_count,
        "Unmatched Visitors": total_visitors - matched_count,
        "Missing Checkouts" : len(missing_co),
        "Unreturned Cards"  : len(unreturned_df),
        # Top lists stored separately — rendered as sub-tables in the workbook
        "_top_hosts"        : top_hosts,
        "_top_companies"    : top_companies,
    }

    return {
        "report_df"    : report_df,
        "discrepancies": disc_df,
        "dashboard"    : dashboard,
        "unreturned_df": unreturned_df,
    }


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT BUILDER
# ──────────────────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=BLACK, size=10) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size)

def _border() -> Border:
    side = Side(style="thin", color=BLACK)
    return Border(left=side, right=side, top=side, bottom=side)

def _style_header(cell, bg=BLACK, fg=WHITE) -> None:
    cell.fill      = _fill(bg)
    cell.font      = _font(bold=True, color=fg, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border    = _border()

def _style_data(cell, bg=WHITE) -> None:
    cell.fill      = _fill(bg)
    cell.font      = _font(size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = _border()

def _auto_col_width(ws, padding: int = 4) -> None:
    """Set each column width to fit its longest value (capped at 48)."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len    = max(
            (len(str(cell.value or "")) for cell in col_cells),
            default=0,
        )
        ws.column_dimensions[col_letter].width = min(max_len + padding, 48)


def build_workbook(
    report_df:     pd.DataFrame,
    disc_df:       pd.DataFrame,
    dashboard:     dict,
    unreturned_df: pd.DataFrame,
) -> bytes:
    """
    Construct the three-sheet output workbook and return it as raw bytes.

    Sheet layout:
        Final Report   – completed visitor log with card, times and duration
        Discrepancies  – all flagged issues grouped by category
        Dashboard      – KPI summary table + unreturned card detail
    """
    wb = Workbook()

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 1 – Final Report
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title          = "Final Report"
    ws.freeze_panes   = "A2"   # keep header row visible while scrolling

    columns = [VR_DATE, VR_CARD, VR_CHECKIN, VR_CHECKOUT, VR_DURATION,
               VR_VISITOR, VR_EMAIL, VR_HOST, VR_COMPANY, VR_PURPOSE]

    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        _style_header(ws.cell(row=1, column=col_idx, value=col_name),
                      bg=BLACK, fg=YELLOW)

    # Data rows with alternating row shading
    for row_idx, (_, row) in enumerate(report_df.iterrows(), start=2):
        row_bg = LIGHT_Y if row_idx % 2 == 0 else WHITE
        for col_idx, col_name in enumerate(columns, start=1):
            val = row.get(col_name, "")
            if col_name == VR_DATE and pd.notna(val):
                try:
                    val = pd.Timestamp(val).strftime("%d-%b-%y")
                except Exception:
                    pass
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _style_data(cell, bg=row_bg)

    _auto_col_width(ws)

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 2 – Discrepancies
    # ─────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Discrepancies")
    ws2.freeze_panes = "A2"

    disc_cols = ["Category", "Date", "Visitor Name", "Card No.", "Detail"]
    for col_idx, col_name in enumerate(disc_cols, start=1):
        _style_header(ws2.cell(row=1, column=col_idx, value=col_name),
                      bg=BLACK, fg=YELLOW)

    if disc_df.empty:
        cell = ws2.cell(row=2, column=1, value="✅  No discrepancies found")
        cell.font = _font(bold=True, color="007700", size=11)
    else:
        for row_idx, (_, row) in enumerate(disc_df.iterrows(), start=2):
            row_bg = LIGHT_Y if row_idx % 2 == 0 else WHITE
            for col_idx, col_name in enumerate(disc_cols, start=1):
                val = row.get(col_name, "")
                # Format date column consistently
                if col_name == "Date" and pd.notna(val) and val != "":
                    try:
                        val = pd.Timestamp(val).strftime("%d-%b-%y")
                    except Exception:
                        pass
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                _style_data(cell, bg=row_bg)
                # Highlight the Category column for visual scanning
                if col_name == "Category":
                    cell.fill = _fill(RED_LT)
                    cell.font = _font(bold=True, size=10)

    _auto_col_width(ws2)

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 3 – Dashboard
    # ─────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Dashboard")
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 22

    # ── Title banner ──────────────────────────────────────────────────────────
    ws3.merge_cells("A1:B1")
    title_cell             = ws3["A1"]
    title_cell.value       = "The Circle — Monthly Visitor Dashboard"
    title_cell.fill        = _fill(YELLOW)
    title_cell.font        = _font(bold=True, color=BLACK, size=14)
    title_cell.alignment   = Alignment(horizontal="center", vertical="center")
    title_cell.border      = _border()
    ws3.row_dimensions[1].height = 32

    # ── KPI summary table ─────────────────────────────────────────────────────
    for col_idx, label in enumerate(["Metric", "Value"], start=1):
        _style_header(ws3.cell(row=2, column=col_idx, value=label),
                      bg=BLACK, fg=YELLOW)

    # Public KPIs (exclude private _top_* keys)
    kpi_items = [
        "Total Visitors",
        "Unique Visitors",
        "Matched Visitors",
        "Unmatched Visitors",
        "Missing Checkouts",
        "Unreturned Cards",
    ]
    for row_idx, key in enumerate(kpi_items, start=3):
        row_bg = LIGHT_Y if row_idx % 2 == 1 else WHITE

        lbl = ws3.cell(row=row_idx, column=1, value=key)
        lbl.fill      = _fill(row_bg)
        lbl.font      = _font(bold=True, size=11)
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        lbl.border    = _border()

        val = ws3.cell(row=row_idx, column=2, value=dashboard[key])
        val.fill      = _fill(YELLOW)
        val.font      = _font(bold=True, color=BLACK, size=14)
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.border    = _border()
        ws3.row_dimensions[row_idx].height = 26

    # ── Helper: write a sub-table starting at given row ───────────────────────
    def _write_subtable(start_row: int, title: str,
                        col_headers: list[str], data: dict) -> int:
        """
        Write a titled two-column sub-table.
        Returns the next available row number after the table.
        """
        # Section title
        ws3.merge_cells(f"A{start_row}:B{start_row}")
        tc             = ws3[f"A{start_row}"]
        tc.value       = title
        tc.fill        = _fill(BLACK)
        tc.font        = _font(bold=True, color=YELLOW, size=11)
        tc.alignment   = Alignment(horizontal="center", vertical="center")
        tc.border      = _border()

        # Column headers
        for ci, h in enumerate(col_headers, start=1):
            _style_header(ws3.cell(row=start_row + 1, column=ci, value=h),
                          bg=BLACK, fg=YELLOW)

        if not data:
            cell      = ws3.cell(row=start_row + 2, column=1, value="No data")
            cell.font = _font(color="888888")
            return start_row + 4

        for ri, (name, count) in enumerate(data.items(), start=start_row + 2):
            row_bg = LIGHT_Y if ri % 2 == 0 else WHITE
            for ci, val in enumerate([name, count], start=1):
                cell           = ws3.cell(row=ri, column=ci, value=val)
                cell.fill      = _fill(row_bg)
                cell.font      = _font(size=10)
                cell.border    = _border()
                cell.alignment = Alignment(horizontal="left" if ci == 1 else "center",
                                           vertical="center")

        return start_row + 2 + len(data) + 1   # +1 blank gap

    # ── Top Hosts sub-table ───────────────────────────────────────────────────
    next_row = len(kpi_items) + 5
    next_row = _write_subtable(
        next_row,
        "Top 5 Hosts by Visitor Count",
        ["Host Name", "Visitors"],
        dashboard.get("_top_hosts", {}),
    )

    # ── Top Companies sub-table ───────────────────────────────────────────────
    next_row += 1   # blank row between sub-tables
    next_row = _write_subtable(
        next_row,
        "Top 5 Visitor Companies",
        ["Company", "Visitors"],
        dashboard.get("_top_companies", {}),
    )

    # ── Unreturned Cards detail sub-table ─────────────────────────────────────
    next_row += 1
    ws3.merge_cells(f"A{next_row}:B{next_row}")
    uc_title             = ws3[f"A{next_row}"]
    uc_title.value       = "Unreturned Cards Detail"
    uc_title.fill        = _fill(BLACK)
    uc_title.font        = _font(bold=True, color=YELLOW, size=11)
    uc_title.alignment   = Alignment(horizontal="center", vertical="center")
    uc_title.border      = _border()

    for col_idx, h in enumerate(["Card No.", "Name"], start=1):
        _style_header(ws3.cell(row=next_row + 1, column=col_idx, value=h),
                      bg=BLACK, fg=YELLOW)

    if unreturned_df.empty:
        cell      = ws3.cell(row=next_row + 2, column=1, value="All cards returned ✅")
        cell.font = _font(bold=True, color="007700")
    else:
        for r_idx, (_, row) in enumerate(unreturned_df.iterrows(),
                                          start=next_row + 2):
            row_bg = LIGHT_Y if r_idx % 2 == 0 else WHITE
            for col_idx, col_key in enumerate([CA_CARDNUM, CA_NAME], start=1):
                val = (int(row[col_key]) if col_key == CA_CARDNUM
                       else str(row[col_key]).strip())
                cell           = ws3.cell(row=r_idx, column=col_idx, value=val)
                cell.fill      = _fill(row_bg)
                cell.font      = _font(size=10)
                cell.border    = _border()
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Serialise to bytes ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ──────────────────────────────────────────────────────────────────────────────
# STREAMLIT UI
# ──────────────────────────────────────────────────────────────────────────────

uploaded_file = st.file_uploader(
    "📂  Upload Visitor Workbook (.xlsx)",
    type=["xlsx"],
    help=(
        "The workbook must contain three sheets: "
        f"'{SHEET_VISITOR}', '{SHEET_CARD}', and '{SHEET_ACCESS}'."
    ),
)

if uploaded_file is not None:
    st.success(f"✅  Workbook uploaded: **{uploaded_file.name}**")
    st.markdown("---")

    col_btn, _ = st.columns([1, 5])
    with col_btn:
        generate = st.button("⚙️  Generate Report", use_container_width=True)

    if generate:
        with st.spinner("Reconciling visitor data…"):
            try:
                xl = pd.ExcelFile(uploaded_file)

                # ── Validate ──────────────────────────────────────────────
                errors = validate_workbook(xl)
                if errors:
                    st.error("**Workbook validation failed:**")
                    for err in errors:
                        st.error(f"• {err}")
                    st.stop()

                # ── Reconcile ─────────────────────────────────────────────
                result        = reconcile(xl)
                report_df     = result["report_df"]
                disc_df       = result["discrepancies"]
                dashboard     = result["dashboard"]
                unreturned_df = result["unreturned_df"]

                # ── KPI Dashboard ─────────────────────────────────────────
                st.markdown(
                    '<div class="sec-head">📊 Reconciliation Summary</div>',
                    unsafe_allow_html=True,
                )
                # Only show public KPI keys (exclude private _top_* keys)
                kpi_keys = [k for k in dashboard.keys() if not k.startswith("_")]
                cols     = st.columns(len(kpi_keys))
                for col, key in zip(cols, kpi_keys):
                    with col:
                        st.markdown(
                            f'<div class="metric-card">'
                            f'<div class="val">{dashboard[key]}</div>'
                            f'<div class="lbl">{key}</div>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

                st.markdown("---")

                # ── Final Report Preview ──────────────────────────────────
                st.markdown(
                    '<div class="sec-head">📋 Final Report Preview</div>',
                    unsafe_allow_html=True,
                )
                preview = report_df.copy()
                if VR_DATE in preview.columns:
                    preview[VR_DATE] = (
                        pd.to_datetime(preview[VR_DATE], errors="coerce")
                        .dt.strftime("%d-%b-%y")
                    )
                st.dataframe(preview, use_container_width=True, height=340)

                # ── Discrepancies Preview ─────────────────────────────────
                st.markdown(
                    '<div class="sec-head">⚠️ Discrepancies</div>',
                    unsafe_allow_html=True,
                )
                if disc_df.empty:
                    st.success("🎉  No discrepancies detected!")
                else:
                    st.dataframe(disc_df, use_container_width=True, height=240)

                st.markdown("---")

                # ── Build & offer download ────────────────────────────────
                xlsx_bytes = build_workbook(
                    report_df, disc_df, dashboard, unreturned_df
                )
                today_str = datetime.now().strftime("%Y-%m-%d")
                filename  = f"Circle_Visitor_Report_{today_str}.xlsx"

                st.markdown(
                    '<div class="sec-head">⬇️ Download Completed Report</div>',
                    unsafe_allow_html=True,
                )
                st.download_button(
                    label    = "⬇️  Download Final Report",
                    data     = xlsx_bytes,
                    file_name= filename,
                    mime     = (
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet"
                    ),
                )

            except Exception as exc:
                st.error(f"**Unexpected error:** {exc}")
                raise   # surface full traceback in terminal for debugging

else:
    st.info("👆  Upload an Excel workbook above to begin.")

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    "<p style='text-align:center;color:#999;font-size:0.8rem;'>"
    "The Circle Visitor Reconciliation Tool &nbsp;·&nbsp; "
    "Built for The Circle.Work"
    "</p>",
    unsafe_allow_html=True,
)
